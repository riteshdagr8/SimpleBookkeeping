"""Build obligations_matrix.xlsx — how the app generates obligations by entity/jurisdiction.

Source of truth: src/lib/obligation-matrix.ts, src/lib/services/obligations.ts
(generateObligationsForClient) and src/lib/compliance-rules.ts (as of 2026-08-16).
"""
import os
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.environ.get("OBLIGATIONS_MATRIX_OUT", "obligations_matrix.xlsx")

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name=FONT, size=10)
MUTED_FONT = Font(name=FONT, size=10, color="7F7F7F")
BOLD_FONT = Font(name=FONT, bold=True, size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="C9C9C9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ZERO_FILL = PatternFill("solid", fgColor="FDE9E9")


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        cell.border = BORDER


def write_table(ws, headers, rows, widths, start_row=1):
    for j, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=j, value=h)
    style_header(ws, len(headers), start_row)
    for i, r in enumerate(rows, start_row + 1):
        for j, v in enumerate(r, 1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = BORDER
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    return start_row + len(rows)


def note(ws, row, text):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = MUTED_FONT
    cell.alignment = WRAP
    return row + 1


wb = Workbook()

# ---------------- Sheet 1: Overview ----------------
ws = wb.active
ws.title = "Overview"
ws["A1"] = "SimpleBookkeeping — How obligations are generated (multi-entity & jurisdiction)"
ws["A1"].font = Font(name=FONT, bold=True, size=15, color="1F4E79")
ws["A2"] = "Source: src/lib/obligation-matrix.ts  +  src/lib/services/obligations.ts  +  src/lib/compliance-rules.ts  (2026-08-16)"
ws["A2"].font = MUTED_FONT

r = 4
ws.cell(row=r, column=1, value="Rules that apply to every obligation type:").font = BOLD_FONT
r += 1
for line in [
    "1. The client's historical review must be marked complete, or NO obligations are generated.",
    "2. Every row is gated on its due date (or its period) falling inside the rolling 12-month window (start of the current month).",
    "3. Re-generation is idempotent (dedup by type + period + due) and purges auto-generated, Pending, future rows whose filing type is no longer valid for the entity/jurisdiction. Historical and completed rows are NEVER deleted.",
    "4. Editing a client's entity type or jurisdiction asks for confirmation before purging invalid future-pending rows.",
    "5. Entity types: Corporation, Self-Employed, Trust, Individual, Partnership. Jurisdictions: Federal + 13 provinces/territories (2-letter codes).",
]:
    ws.cell(row=r, column=1, value=line).font = BODY_FONT
    r += 1

r += 1
hdr = ["Category", "Entity / jurisdiction trigger", "Filing type(s)", "Filing due", "Payment due"]
rows = [
    ["Income tax", "Corporation", "T2", "FYE + 6 mo", "FYE + 2 mo (3 if threeMonthEligible)"],
    ["Income tax", "Self-Employed", "T1", "Jun 15 (year after FYE)", "Apr 30 (year after FYE)"],
    ["Income tax", "Individual", "T1", "Apr 30 (year after FYE)", "Apr 30 (year after FYE)"],
    ["Income tax", "Partnership", "T5013", "Mar 31 (year after FYE)", "N/A"],
    ["Income tax", "Trust", "T3", "FYE + 90 d", "FYE + 90 d"],
    ["Sales tax", "Harmonized (ON, NB, NS, NL, PE)", "HST", "by frequency (see Sales Tax sheet)", "= filing due"],
    ["Sales tax", "Non-PST (AB, NT, NU, YT)", "GST", "by frequency", "= filing due"],
    ["Sales tax", "Quebec", "GST/QST", "by frequency", "= filing due"],
    ["Sales tax", "BC / SK", "GST + PST", "by frequency", "= filing due"],
    ["Sales tax", "Manitoba", "GST + RST", "by frequency", "= filing due"],
    ["Annual return", "Federal jurisdiction + incorporation date", "FederalAnnualReturn", "incorporation anniversary + 60 d", "-"],
    ["Annual return", "Any province/territory (Corporation)", "ProvincialAnnualReturn", "by jurisdiction (see Annual Returns sheet)", "-"],
    ["Payroll", "payrollApplicable + remitterType", "PayrollRemittance", "Monthly: 15th of following month; Quarterly: 15th after quarter-end", "= filing due"],
    ["Payroll", "payrollApplicable + payrollFrequency", "PayrollProcessing", "= payroll period end", "= period end"],
    ["Info returns", "Corporation", "T4, T4A, T5", "last day of Feb (FYE year + 1)", "-"],
    ["Info returns", "Self-Employed / Individual (if payrollApplicable)", "T4, T4A", "last day of Feb (FYE year + 1)", "-"],
    ["Info returns", "Partnership (if payrollApplicable)", "T4, T4A, T5", "last day of Feb (FYE year + 1)", "-"],
    ["Info returns", "Trust", "T3Slips", "FYE + 90 d", "-"],
]
r = write_table(ws, hdr, rows, [16, 34, 26, 34, 30], start_row=r)
r += 1
r = note(ws, r, "Sales tax is gated on hstApplicable=ON + hstFrequency; jurisdiction selects which filings. Annual returns are suppressed for Self-Employed/Trust/Individual/Partnership. Sales tax, payroll, and info returns for Federal jurisdiction use the HST / standard defaults.")

# ---------------- Sheet 2: Income Tax ----------------
ws = wb.create_sheet("Income Tax")
ws["A1"] = "Income taxes by entity type (exactly one per client)"
ws["A1"].font = Font(name=FONT, bold=True, size=14, color="1F4E79")
hdr = ["Entity type", "Filing type", "Period", "Filing due", "Payment due"]
rows = [
    ["Corporation", "T2", "prior FYE -> FYE", "FYE + 6 months", "FYE + 2 months (3 if threeMonthEligible)"],
    ["Self-Employed", "T1", "Jan 1 -> Dec 31 (FYE year)", "Jun 15 of following year", "Apr 30 of following year"],
    ["Individual", "T1", "Jan 1 -> Dec 31 (FYE year)", "Apr 30 of following year", "Apr 30 of following year"],
    ["Partnership", "T5013", "Jan 1 -> Dec 31 (FYE year)", "Mar 31 of following year", "N/A"],
    ["Trust", "T3", "prior FYE -> FYE", "FYE + 90 days", "FYE + 90 days"],
]
r = write_table(ws, hdr, rows, [18, 12, 28, 30, 34], start_row=3)
r += 1
r = note(ws, r, "Legacy clients with no entity type are treated as Corporation (T2).")

# ---------------- Sheet 3: Sales Tax ----------------
ws = wb.create_sheet("Sales Tax")
ws["A1"] = "Sales tax by jurisdiction and frequency"
ws["A1"].font = Font(name=FONT, bold=True, size=14, color="1F4E79")
hdr = ["Jurisdiction", "Filings", "Frequency", "Filing / payment due"]
rows = [
    ["ON, NB, NS, NL, PE", "HST", "Monthly", "last day of the following month"],
    ["ON, NB, NS, NL, PE", "HST", "Quarterly", "last day of the month after quarter-end"],
    ["ON, NB, NS, NL, PE", "HST", "Annual (Corporation)", "FYE + 3 months"],
    ["AB, NT, NU, YT", "GST", "any", "same due rules as above"],
    ["QC", "GST/QST (combined)", "any", "same due rules as above"],
    ["BC, SK", "GST + PST", "any", "same due rules as above"],
    ["MB", "GST + RST", "any", "same due rules as above"],
    ["Self-Employed / Individual with Dec-31 FYE", "any", "Annual", "payment Apr 30; filing Jun 15 of following year"],
]
r = write_table(ws, hdr, rows, [34, 24, 24, 42], start_row=3)
r += 1
r = note(ws, r, "Monthly is anchored to the rolling window; Quarterly/Annual are anchored to gstYearEnd (or calendar quarters) and the FYE year.")

# ---------------- Sheet 4: Annual Returns ----------------
ws = wb.create_sheet("Annual Returns")
ws["A1"] = "Corporate annual returns (only for Corporation entities)"
ws["A1"].font = Font(name=FONT, bold=True, size=14, color="1F4E79")
hdr = ["Jurisdiction", "Filing type", "Deadline"]
rows = [
    ["Federal", "FederalAnnualReturn", "incorporation anniversary + 60 days (period: due-60d -> due)"],
    ["ON, QC", "ProvincialAnnualReturn", "FYE + 6 months (period: prior FYE -> FYE)"],
    ["BC, NS, NL", "ProvincialAnnualReturn", "incorporation anniversary + 60 days"],
    ["AB, SK, MB, NB, YT, NT, NU", "ProvincialAnnualReturn", "last day of the month following the anniversary month"],
    ["PE", "ProvincialAnnualReturn", "last day of the anniversary month"],
]
r = write_table(ws, hdr, rows, [34, 24, 56], start_row=3)
r += 1
r = note(ws, r, "Anniversary-based provinces require an incorporation date. Suppressed entirely for Self-Employed, Trust, Individual, and Partnership entities.")

# ---------------- Sheet 5: Payroll ----------------
ws = wb.create_sheet("Payroll")
ws["A1"] = "Payroll"
ws["A1"].font = Font(name=FONT, bold=True, size=14, color="1F4E79")

ws["A3"] = "PayrollRemittance — by remitterType"
ws["A3"].font = BOLD_FONT
hdr = ["remitterType", "Rows", "Period", "Filing / payment due"]
rows = [
    ["Monthly", 12, "monthly, 1st -> last day", "15th of the month after the period"],
    ["Quarterly", 4, "calendar quarters", "15th of the month after the quarter end"],
]
r = write_table(ws, hdr, rows, [16, 8, 30, 34], start_row=4)

r = ws.max_row + 3
ws.cell(row=r, column=1, value="PayrollProcessing — by payrollFrequency").font = BOLD_FONT
r += 1
hdr = ["payrollFrequency", "Rows", "Period", "Filing / payment due"]
rows = [
    ["Weekly", 52, "7-day runs", "= period end"],
    ["Bi-Weekly", 26, "14-day runs", "= period end"],
    ["Semi-Monthly", 24, "1st-15th, 16th-last day", "= period end"],
    ["Monthly", 12, "1st -> last day", "= period end"],
]
write_table(ws, hdr, rows, [16, 8, 30, 22], start_row=r)

# ---------------- Sheet 6: Info Returns ----------------
ws = wb.create_sheet("Info Returns")
ws["A1"] = "Information returns by entity type"
ws["A1"].font = Font(name=FONT, bold=True, size=14, color="1F4E79")
hdr = ["Entity type", "Filings", "Filing due"]
rows = [
    ["Corporation", "T4, T4A, T5", "last day of February (FYE year + 1)"],
    ["Self-Employed (if payrollApplicable)", "T4, T4A", "last day of February (FYE year + 1)"],
    ["Individual (if payrollApplicable)", "T4, T4A", "last day of February (FYE year + 1)"],
    ["Partnership (if payrollApplicable)", "T4, T4A, T5", "last day of February (FYE year + 1)"],
    ["Trust", "T3Slips", "FYE + 90 days"],
]
r = write_table(ws, hdr, rows, [34, 22, 40], start_row=3)

# ---------------- Sheet 7: Examples ----------------
ws = wb.create_sheet("Examples")
ws["A1"] = "Worked examples (illustrative) — what a generated schedule looks like"
ws["A1"].font = Font(name=FONT, bold=True, size=14, color="1F4E79")
hdr = ["Client config", "Expected obligations (rolling window)"]
examples = [
    ["Corporation, Ontario, sales tax Monthly", "T2, HST x 12, ProvincialAnnualReturn, T4, T4A, T5"],
    ["Corporation, BC, sales tax Monthly", "T2, GST x 12, PST x 12, ProvincialAnnualReturn, T4, T4A, T5"],
    ["Corporation, Quebec, sales tax Quarterly", "T2, GST/QST x 4, ProvincialAnnualReturn, T4, T4A, T5"],
    ["Corporation, Federal, no sales tax", "T2, FederalAnnualReturn, T4, T4A, T5"],
    ["Trust, Ontario", "T3, T3Slips (no T2 / annual returns / T4-T5)"],
    ["Self-Employed, Ontario, sales tax Annual", "T1 (payment Apr 30, filing Jun 15), HST"],
    ["Individual, Ontario", "T1 (filing & payment Apr 30)"],
    ["Partnership, Ontario, payroll", "T5013, PayrollRemittance x 12, PayrollProcessing x 12, T4, T4A, T5"],
]
r = write_table(ws, hdr, examples, [40, 70], start_row=3)
r += 1
r = note(ws, r, "A row is only included when its due date/period lands inside the current rolling 12-month window, so exact counts vary with today's date.")

wb.save(OUT)
print(f"saved {OUT}")
